import cv2
import numpy as np
import face_recognition
import os
from datetime import datetime, date
import openpyxl
from tkinter import *
from tkinter import messagebox
import pyttsx3
import time

path = 'Images'
today_date = date.today().strftime("%Y-%m-%d")
attendance_file = f'Attendance_{today_date}.xlsx'

if not os.path.exists(path):
    os.makedirs(path)


if not os.path.exists(attendance_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"
    sheet.append(["Name", "Date", "Time", "Status"])
    workbook.save(attendance_file)

engine = pyttsx3.init()
engine.setProperty('rate', 170)
engine.setProperty('volume', 1.0)

def speak(text):
    """Speak given text aloud"""
    engine.say(text)
    engine.runAndWait()


def load_images():
    images = []
    classNames = []
    myList = os.listdir(path)
    for imgname in myList:
        curImg = cv2.imread(f'{path}/{imgname}')
        if curImg is None:
            print(f"⚠️ Unable to read image: {imgname}")
            continue
        images.append(curImg)
        classNames.append(os.path.splitext(imgname)[0])
    return images, classNames

def findEncodings(images):
    encodeList = []
    for img in images:
        try:
            img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            encodes = face_recognition.face_encodings(img)
            if len(encodes) > 0:
                encodeList.append(encodes[0])
            else:
                print(" No face found in one image — skipping it.")
        except Exception as e:
            print(f" Error processing image: {e}")
    return encodeList

#
def markAttendance(name, status="Present"):
    workbook = openpyxl.load_workbook(attendance_file)
    sheet = workbook.active
    today = date.today().strftime("%Y-%m-%d")
    now = datetime.now().strftime("%H:%M:%S")

    
    for row in sheet.iter_rows(values_only=True):
        if row[0] == name and row[1] == today:
            return

    sheet.append([name, today, now, status])
    workbook.save(attendance_file)
    print(f"{name} marked as {status} ✅")
    speak(f"{name} is {status}")


def start_recognition():
    images, classNames = load_images()
    if len(images) == 0:
        messagebox.showerror("Error", "No student images found in 'Images' folder.")
        return

    encodeListKnown = findEncodings(images)
    cap = cv2.VideoCapture(0)

    if not cap.isOpened():
        messagebox.showerror("Camera Error", "Could not access webcam.")
        return

    present_students = set()
    start_time = time.time()

    messagebox.showinfo("Info", "Camera started! Press 'Q' to stop early or wait 15 seconds.")

    while True:
        success, img = cap.read()
        if not success:
            break

        imgS = cv2.resize(img, (0, 0), None, 0.5, 0.5)
        imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

        facesCurFrame = face_recognition.face_locations(imgS)
        encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)

        for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
            matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
            faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
            matchIndex = np.argmin(faceDis)

            if matches[matchIndex]:
                name = classNames[matchIndex].upper()
                present_students.add(name)
                markAttendance(name, "Present")

                y1, x2, y2, x1 = faceLoc
                y1, x2, y2, x1 = y1 * 2, x2 * 2, y2 * 2, x1 * 2
                cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                cv2.putText(img, name, (x1 + 6, y2 - 6),
                            cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)

        cv2.imshow('Face Recognition Attendance', img)

       
        if time.time() - start_time > 15:
            break
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    
    for student in classNames:
        student = student.upper()
        if student not in present_students:
            markAttendance(student, "Absent")

    cap.release()
    cv2.destroyAllWindows()
    messagebox.showinfo("Attendance Complete", f"Attendance saved in {attendance_file}")

def register_student():
    name = name_entry.get().strip()
    if name == "":
        messagebox.showerror("Error", "Please enter a student name!")
        return

    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        messagebox.showerror("Error", "Cannot open camera.")
        return

    messagebox.showinfo("Info", "Press 'Spacebar' to capture the face, 'Q' to quit.")
    speak(f"Capturing face for {name}")

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        cv2.imshow('Register New Student', frame)
        key = cv2.waitKey(1)

        # Press Spacebar to capture image
        if key == 32:  # SPACE key
            file_path = os.path.join(path, f"{name}.jpg")
            cv2.imwrite(file_path, frame)
            messagebox.showinfo("Saved", f"Image saved as {file_path}")
            speak(f"Face saved for {name}")
            break

        elif key == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()


def mark_present():
    name = name_entry.get().strip()
    if name == "":
        messagebox.showerror("Error", "Enter name to mark Present")
        return
    markAttendance(name, "Present")
    messagebox.showinfo("Success", f"{name} marked as Present")

def mark_leave():
    name = name_entry.get().strip()
    if name == "":
        messagebox.showerror("Error", "Enter name to mark Leave")
        return
    markAttendance(name, "Leave")
    messagebox.showinfo("Success", f"{name} marked on Leave")

def open_attendance():
    os.startfile(attendance_file)


root = Tk()
root.title("Face Recognition Attendance System")
root.geometry("600x530")
root.config(bg="#e6f0ff")

Label(root, text="FACE ATTENDANCE SYSTEM", font=("Arial", 20, "bold"),
      bg="#007acc", fg="white").pack(fill=X, pady=10)

Label(root, text="Enter Name:", font=("Arial", 12),
      bg="#e6f0ff").pack(pady=5)
name_entry = Entry(root, font=("Arial", 12), width=30)
name_entry.pack(pady=5)

Button(root, text=" Register New Student", font=("Arial", 12), bg="#ff9800", fg="white",
       command=register_student).pack(pady=10)

Button(root, text=" Start Recognition", font=("Arial", 12), bg="#28a745", fg="white",
       command=start_recognition).pack(pady=10)

Button(root, text=" Mark Present", font=("Arial", 12), bg="#17a2b8", fg="white",
       command=mark_present).pack(pady=5)

Button(root, text="🗓 Mark Leave", font=("Arial", 12), bg="#ffc107", fg="black",
       command=mark_leave).pack(pady=5)

Button(root, text=" View Today's Attendance", font=("Arial", 12), bg="#007bff", fg="white",
       command=open_attendance).pack(pady=10)

Button(root, text=" Exit", font=("Arial", 12), bg="gray", fg="white",
       command=root.destroy).pack(pady=5)

root.mainloop()
